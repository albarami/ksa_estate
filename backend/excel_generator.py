"""Excel generator replicating Al-Hada Opportunity.xlsx layout exactly.

Sheet 1: Assumptions — exact Al-Hada replica with LIVE formulas
Sheet 2: Zoning Report — parcel + decoded regulations
Sheet 3: Market Data — SREM intelligence
Sheet 4: Sensitivity Analysis — 5x5 IRR matrix
Sheet 5: Scenario Comparison — 3 scenarios side-by-side
"""

from __future__ import annotations

import io
import re
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from backend.excel_labels import get_labels as _L


# ---------------------------------------------------------------------------
# Al-Hada style constants
# ---------------------------------------------------------------------------

FONT_AR = "Sakkal Majalla"
F_NORMAL = Font(name=FONT_AR, size=12)
F_BOLD = Font(name=FONT_AR, size=12, bold=True)
F_HEADER = Font(name=FONT_AR, size=14, bold=True)
F_TOTAL = Font(name=FONT_AR, size=12, bold=True, color="FFFFFF")
F_KPI = Font(name=FONT_AR, size=14, bold=True)
F_SMALL = Font(name=FONT_AR, size=10)
F_LINK = Font(name=FONT_AR, size=10, color="0563C1", underline="single")

FILL_DATA = PatternFill(start_color="FAFBF4", end_color="FAFBF4", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="898989", end_color="898989", fill_type="solid")
FILL_SECTION = PatternFill(start_color="969183", end_color="969183", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
FILL_RED = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
FILL_HEADER = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT_AL = Alignment(horizontal="right", vertical="center")

THIN = Side(style="thin")
MED = Side(style="medium")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOP_MED = Border(top=MED)

SAR_FMT = '_(* #,##0_);_(* \\(#,##0\\);_(* "-"??_);_(@_)'
SAR2_FMT = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'
PCT_FMT = "0%"
PCT2_FMT = "0.00%"
PCTD_FMT = "0.0%"
NUM_FMT = '#,##0_);[Red](#,##0)'


def _c(ws, row, col, value=None, font=None, fill=None, fmt=None, align=None, border=None):
    """Write a cell with styling."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


# ---------------------------------------------------------------------------
# Sheet 1: Assumptions (Al-Hada exact replica)
# ---------------------------------------------------------------------------

def _build_assumptions_sheet(wb: Workbook, pf: dict, land: dict, L: dict, rtl: bool) -> None:
    """Al-Hada style Assumptions sheet with LIVE Excel formulas throughout."""
    ws = wb.active
    ws.title = L["assumptions"]
    ws.sheet_view.rightToLeft = rtl

    widths = {"B": 23, "C": 3, "D": 39, "E": 15, "F": 15, "G": 15, "H": 17, "I": 5,
              "J": 3, "K": 3, "L": 26, "M": 10, "N": 20, "O": 20, "P": 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    fs = pf.get("fund_size", {})
    lc = pf.get("land_costs", {})
    cc = pf.get("construction_costs", {})
    fin = pf.get("financing", {})
    cf = pf.get("cash_flows", {})
    inputs = pf.get("inputs_used", {})

    def iv(key):
        return inputs.get(key, {}).get("value")

    def cf_val(key: str, idx: int) -> float:
        arr = cf.get(key, [])
        return arr[idx] if idx < len(arr) else 0

    n = int(iv("fund_period_years") or 3)
    years = list(range(1, n + 1))
    far_val = iv("far") or 1.0
    ltv = iv("bank_ltv_pct") or 0.667
    resolved_area = iv("land_area_sqm") or land.get("area_sqm", 0)
    in_kind_pct = lc.get("in_kind_pct", 0) or 0

    # =====================================================================
    # LEFT SIDE (cols B-I): Fund header, Land, Construction, Sales, Fees
    # =====================================================================

    # --- Rows 2-7: Fund header ---
    _c(ws, 2, 4, "Google Maps Link", F_SMALL)
    _c(ws, 2, 5, f"Parcel {land.get('parcel_id')} - {land.get('district_name')}", F_BOLD)
    ws.merge_cells("E2:F2")
    _c(ws, 3, 4, L["fund_name"], F_BOLD)
    _c(ws, 3, 5, L["fund_type_val"], F_BOLD)
    ws.merge_cells("E3:F3")
    _c(ws, 4, 4, L["fund_type_label"], F_BOLD)
    _c(ws, 4, 5, L["fund_type_val"], F_BOLD)
    ws.merge_cells("E4:F4")
    _c(ws, 5, 4, L["fund_period"], F_BOLD)
    _c(ws, 5, 5, n, F_BOLD, fmt="#,##0")
    ws.merge_cells("E5:F5")
    _c(ws, 6, 4, L["fund_size_label"], F_BOLD)
    _c(ws, 6, 5, "=F54", F_BOLD, fmt=SAR_FMT)
    ws.merge_cells("E6:F6")
    _c(ws, 7, 4, L["total_equity"], F_BOLD)
    _c(ws, 7, 5, "=F57", F_BOLD, fmt=SAR_FMT)
    ws.merge_cells("E7:F7")

    # --- Row 9-16: Land section ---
    _c(ws, 9, 2, L["land_section"], F_HEADER, FILL_SECTION)
    _c(ws, 9, 4, L["land_assumptions"], F_HEADER)
    _c(ws, 10, 6, L.get("total_label", "Total"), F_BOLD, align=CENTER)
    _c(ws, 10, 8, L.get("cash", "Cash"), F_BOLD, align=CENTER)
    _c(ws, 10, 9, L.get("inkind", "In-Kind"), F_BOLD, align=CENTER)

    _c(ws, 11, 4, L["land_area"], F_NORMAL)
    _c(ws, 11, 6, resolved_area, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 11, 8, 1.0 - in_kind_pct, F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 11, 9, in_kind_pct, F_NORMAL, fmt=PCT_FMT, align=CENTER)

    _c(ws, 12, 4, L["land_price"], F_NORMAL)
    _c(ws, 12, 5, iv("land_price_per_sqm") or 0, F_NORMAL, fmt=SAR2_FMT, align=CENTER)
    _c(ws, 12, 6, "=$E$12*F11", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 12, 8, "=H11*F12", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 12, 9, "=I11*F12", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 13, 4, L["brokerage"], F_NORMAL)
    _c(ws, 13, 5, iv("brokerage_fee_pct") or 0.025, F_NORMAL, fmt=PCTD_FMT, align=CENTER)
    _c(ws, 13, 6, "=$E$13*F12", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 14, 4, L["transfer_tax"], F_NORMAL)
    _c(ws, 14, 5, iv("real_estate_transfer_tax_pct") or 0.05, F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 14, 6, "=IF(I11>0,0,E14*F12)", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 15, 4, L["brokerage_vat"], F_NORMAL)
    _c(ws, 15, 5, iv("brokerage_vat_pct") or 0.15, F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 15, 6, "=E15*F13", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 16, 4, L["total"], F_TOTAL, FILL_TOTAL)
    _c(ws, 16, 6, "=SUM(F12:F15)", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # --- Rows 18-29: Construction ---
    _c(ws, 18, 2, L["costs_section"], F_HEADER, FILL_SECTION)
    _c(ws, 18, 4, L["cost_assumptions"], F_HEADER)
    _c(ws, 19, 4, L.get("direct_costs", "Direct Costs"), F_BOLD)
    _c(ws, 19, 5, "%", F_BOLD, align=CENTER)
    _c(ws, 19, 6, L.get("area_col", "Area"), F_BOLD, align=CENTER)
    _c(ws, 19, 7, L.get("cost_per_m", "Cost/m\u00b2"), F_BOLD, align=CENTER)
    _c(ws, 19, 8, L.get("total_col", "Total"), F_BOLD, align=CENTER)

    # FIX A: GBA as formula =F11*FAR (not hardcoded)
    _c(ws, 20, 4, L.get("infrastructure", "Infrastructure"), F_NORMAL)
    _c(ws, 20, 5, 1, F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 20, 6, f"=F11*{far_val}", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # GBA = area * FAR
    _c(ws, 20, 7, iv("infrastructure_cost_per_sqm") or 500, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 20, 8, "=G20*F20", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 21, 4, L.get("superstructure", "Superstructure"), F_NORMAL)
    _c(ws, 21, 6, "=F20", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 21, 7, iv("superstructure_cost_per_sqm") or 2500, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 21, 8, "=G21*F21", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 22, 4, L.get("parking", "Parking"), F_NORMAL)
    _c(ws, 22, 6, iv("parking_area_sqm") or 0, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 22, 7, iv("parking_cost_per_sqm") or 2000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 22, 8, "=G22*F22", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 23, 4, L.get("total_direct", "Total Direct"), F_TOTAL, FILL_TOTAL)
    _c(ws, 23, 8, "=SUM(H20:H22)", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 24, 4, L.get("indirect_costs", "Indirect Costs"), F_BOLD)
    _c(ws, 24, 5, "%", F_BOLD, align=CENTER)
    _c(ws, 24, 6, L.get("total_col", "Total"), F_BOLD, align=CENTER)

    _c(ws, 25, 4, L.get("developer_fee", "Developer Fee"), F_NORMAL)
    _c(ws, 25, 5, iv("developer_fee_pct") or 0.10, F_NORMAL, fmt=PCTD_FMT, align=CENTER)
    _c(ws, 25, 6, "=E25*H23", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 26, 4, L.get("other_indirect", "Other Indirect"), F_NORMAL)
    _c(ws, 26, 5, iv("other_indirect_pct") or 0.06, F_NORMAL, fmt=PCTD_FMT, align=CENTER)
    _c(ws, 26, 6, "=E26*H23", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 27, 4, L.get("contingency", "Contingency"), F_NORMAL)
    _c(ws, 27, 5, iv("contingency_pct") or 0.05, F_NORMAL, fmt=PCTD_FMT, align=CENTER)
    _c(ws, 27, 6, "=E27*H23", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 28, 4, L.get("total_indirect", "Total Indirect"), F_TOTAL, FILL_TOTAL)
    _c(ws, 28, 8, "=F25+F26+F27", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 29, 4, L.get("total_all_costs", "Total All Costs"), F_TOTAL, FILL_TOTAL)
    _c(ws, 29, 8, "=H28+H23", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # --- Rows 31-33: Sales ---
    _c(ws, 31, 2, L.get("sales_section", "Sales"), F_HEADER, FILL_SECTION)
    _c(ws, 31, 4, L.get("sales_assumptions", "Sales Assumptions"), F_HEADER)
    _c(ws, 31, 6, L.get("area_col", "Area"), F_BOLD, align=CENTER)
    _c(ws, 31, 7, L.get("price_per_m", "Price/m\u00b2"), F_BOLD, align=CENTER)
    _c(ws, 31, 8, L.get("total_sales", "Total Sales"), F_BOLD, align=CENTER)

    _c(ws, 32, 4, L.get("unit_sales", "Unit Sales"), F_NORMAL)
    _c(ws, 32, 6, "=F20", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # GBA
    _c(ws, 32, 7, iv("sale_price_per_sqm") or 0, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 32, 8, "=G32*F32", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 33, 4, L.get("total_label", "Total"), F_TOTAL, FILL_TOTAL)
    _c(ws, 33, 8, "=H32", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # --- Rows 36-52: Fund fees (FIX B: formulas, no circular refs) ---
    _c(ws, 36, 2, L.get("financing_section", "Fund Fees"), F_HEADER, FILL_SECTION)
    _c(ws, 36, 4, L.get("total_financing", "Financing & Fund Cost"), F_HEADER)
    _c(ws, 37, 4, " ", F_BOLD)
    _c(ws, 37, 5, L.get("cost_per_m", "Rate"), F_BOLD, align=CENTER)
    _c(ws, 37, 6, L.get("total_col", "Total (SAR)"), F_BOLD, align=CENTER)

    # Category 1: Based on construction/loan cost (bank loan = H23 * LTV)
    # Interest = bank_loan * rate * years
    _c(ws, 38, 4, L.get("interest", "Interest"), F_NORMAL)
    _c(ws, 38, 5, iv("interest_rate_pct") or 0.08, F_NORMAL, fmt=PCT2_FMT, align=CENTER)
    _c(ws, 38, 6, f"=H23*{ltv}*E38*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    # Arrangement fee = bank_loan * rate (one-time)
    _c(ws, 39, 4, L.get("arrangement_fee", "Arrangement Fee"), F_NORMAL)
    _c(ws, 39, 5, iv("arrangement_fee_pct") or 0.02, F_NORMAL, fmt=PCT2_FMT, align=CENTER)
    _c(ws, 39, 6, f"=H23*{ltv}*E39", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    # Category 2: Fixed annual amounts x period (Al-Hada approach, no circularity)
    # Management fee: annual amount x years
    mgmt_annual = (iv("management_fee_pct") or 0.015) * (fs.get("total_fund_size", 0) or 1) / max(n, 1)
    _c(ws, 40, 4, L.get("mgmt_fee", "Management Fee"), F_NORMAL)
    _c(ws, 40, 5, round(mgmt_annual), F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 40, 6, "=E40*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 41, 4, L.get("custodian", "Custodian"), F_NORMAL)
    _c(ws, 41, 5, iv("custodian_fee_annual") or 50000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 41, 6, "=E41*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 42, 4, L.get("board", "Board"), F_NORMAL)
    _c(ws, 42, 5, iv("board_fee_annual") or 100000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 42, 6, "=E42*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 43, 4, L.get("sharia_cert", "Sharia Certificate"), F_NORMAL)
    _c(ws, 43, 5, iv("sharia_certificate_fee") or 5000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 43, 6, "=E43", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # one-time

    _c(ws, 44, 4, L.get("sharia_board", "Sharia Board"), F_NORMAL)
    _c(ws, 44, 5, iv("sharia_board_fee_annual") or 5000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 44, 6, "=E44*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 45, 4, L.get("legal", "Legal"), F_NORMAL)
    _c(ws, 45, 5, iv("legal_counsel_fee") or 50000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 45, 6, "=E45", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # one-time

    _c(ws, 46, 4, L.get("auditor", "Auditor"), F_NORMAL)
    _c(ws, 46, 5, iv("auditor_fee_annual") or 50000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 46, 6, "=E46*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 47, 4, L.get("valuation", "Valuation"), F_NORMAL)
    _c(ws, 47, 5, iv("valuation_fee_quarterly") or 20000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 47, 6, "=E47*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 48, 4, L.get("reserve", "Other Reserve"), F_NORMAL)
    _c(ws, 48, 5, iv("other_reserve_pct") or 0.0005, F_NORMAL, fmt=PCT2_FMT, align=CENTER)
    _c(ws, 48, 6, "=E48*H23*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # % of construction

    _c(ws, 49, 4, L.get("spv", "SPV Fee"), F_NORMAL)
    _c(ws, 49, 5, iv("spv_formation_fee") or 25000, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 49, 6, "=E49", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # one-time

    _c(ws, 50, 4, L.get("structuring", "Structuring"), F_NORMAL)
    _c(ws, 50, 5, iv("structuring_fee_pct") or 0.01, F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 50, 6, "=H23*E50", F_NORMAL, fmt=SAR_FMT, align=CENTER)  # % of construction

    _c(ws, 51, 4, L.get("operator", "Operator"), F_NORMAL)
    _c(ws, 51, 5, iv("operator_fee_pct") or 0.0015, F_NORMAL, fmt=PCT2_FMT, align=CENTER)
    _c(ws, 51, 6, "=E51*H23*E5", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 52, 4, L.get("total_financing", "Total Financing"), F_TOTAL, FILL_TOTAL)
    _c(ws, 52, 6, "=SUM(F38:F51)", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # --- Row 54: Total Fund Size ---
    _c(ws, 54, 4, L.get("total_fund_size", "Total Fund Size"), F_TOTAL, FILL_TOTAL)
    _c(ws, 54, 6, "=F16+H29+F52", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # --- Rows 56-60: Capital structure ---
    _c(ws, 56, 4, L.get("capital_structure", "Capital Structure"), F_HEADER)
    _c(ws, 57, 4, L.get("equity", "Equity"), F_BOLD)
    _c(ws, 57, 5, "=F57/$F$60", F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 57, 6, "=F54-SUM(F58:F59)", F_BOLD, fmt=SAR_FMT, align=CENTER)

    _c(ws, 58, 4, L.get("inkind_owner", "In-Kind"), F_NORMAL)
    _c(ws, 58, 5, "=F58/$F$60", F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 58, 6, "=I12", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    _c(ws, 59, 4, L.get("bank_financing", "Bank Financing"), F_NORMAL)
    _c(ws, 59, 5, "=F59/$F$60", F_NORMAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 59, 6, f"={ltv}*F12", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 59, 7, ltv, F_NORMAL, fmt=PCT_FMT, align=CENTER)

    _c(ws, 60, 4, L.get("total_capital", "Total Capital"), F_TOTAL, FILL_TOTAL)
    _c(ws, 60, 5, "=SUM(E57:E59)", F_TOTAL, FILL_TOTAL, fmt=PCT_FMT, align=CENTER)
    _c(ws, 60, 6, "=SUM(F57:F59)", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # =====================================================================
    # RIGHT SIDE (cols L-P): Cash flows, Fund structure, KPIs
    # =====================================================================

    _c(ws, 3, 12, L["currency"], F_BOLD)
    _c(ws, 4, 12, L["year_label"], F_BOLD)
    for i in range(len(years)):
        _c(ws, 4, 14 + i, 2025 + i, F_BOLD, align=CENTER)
    _c(ws, 5, 12, " ", F_BOLD)
    for i, yr in enumerate(years):
        _c(ws, 5, 14 + i, f"Y{yr}", F_BOLD, align=CENTER)

    # Cash inflows
    _c(ws, 6, 12, L["cf_inflows"], F_BOLD)
    _c(ws, 7, 12, L.get("cf_sales", "Sales"), F_NORMAL)
    for i in range(len(years)):
        _c(ws, 7, 14 + i, cf_val("inflows_sales", i), F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 8, 12, L.get("total_label", "Total"), F_BOLD)
    for i in range(len(years)):
        col_l = get_column_letter(14 + i)
        _c(ws, 8, 14 + i, f"=SUM({col_l}7)", F_BOLD, fmt=SAR_FMT, align=CENTER)

    # Cash outflows
    _c(ws, 9, 12, L.get("cf_outflows", "Cash Outflows"), F_BOLD)

    # Outflow rows with data from computation engine
    cf_rows = [
        (11, L.get("cf_land", "Land"), "outflows_land"),
        (12, L.get("cf_direct", "Direct Costs"), "outflows_direct"),
        (13, L.get("cf_indirect", "Indirect Costs"), "outflows_indirect"),
        (14, L.get("cf_interest", "Interest"), "outflows_interest"),
        (15, L.get("arrangement_fee", "Arrangement Fee"), None),
        (16, L.get("cf_fees", "Fund Fees"), "outflows_fees"),
    ]
    for row, label, key in cf_rows:
        _c(ws, row, 12, label, F_NORMAL)
        for i in range(len(years)):
            if key:
                val = cf_val(key, i)
            elif row == 15:
                val = fin.get("arrangement_fee", 0) if i == 0 else 0
            else:
                val = 0
            _c(ws, row, 14 + i, val, F_NORMAL, fmt=SAR_FMT, align=CENTER)

    # Total outflows
    _c(ws, 28, 12, L.get("cf_total", "Total Outflows"), F_TOTAL, FILL_TOTAL)
    for i in range(len(years)):
        col_l = get_column_letter(14 + i)
        _c(ws, 28, 14 + i, f"=SUM({col_l}11:{col_l}16)", F_TOTAL, FILL_TOTAL, fmt=SAR_FMT, align=CENTER)

    # Net cash flow
    _c(ws, 30, 12, L.get("cf_net", "Net Cash Flow"), F_BOLD, border=BORDER_TOP_MED)
    for i in range(len(years)):
        col_l = get_column_letter(14 + i)
        _c(ws, 30, 14 + i, f"={col_l}8-{col_l}28", F_BOLD, fmt=SAR_FMT, align=CENTER, border=BORDER_TOP_MED)

    # Cumulative
    _c(ws, 31, 12, L.get("cf_cumulative", "Cumulative"), F_NORMAL)
    _c(ws, 31, 14, "=N30", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    for i in range(1, len(years)):
        col_l = get_column_letter(14 + i)
        prev_l = get_column_letter(13 + i)
        _c(ws, 31, 14 + i, f"={col_l}30+{prev_l}31", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    # Fund capital structure
    _c(ws, 33, 12, L.get("cf_fund_capital", "Fund Capital"), F_BOLD)
    _c(ws, 34, 12, L.get("equity", "Equity"), F_NORMAL)
    _c(ws, 34, 14, "=F57", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 35, 12, L.get("inkind_owner", "In-Kind"), F_NORMAL)
    _c(ws, 35, 14, "=I12", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    bank_loan = fin.get("bank_loan_amount", 0) if isinstance(fin.get("bank_loan_amount"), (int, float)) else 0
    _c(ws, 36, 12, "Debt Withdrawal", F_NORMAL)
    _c(ws, 36, 14, bank_loan, F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 37, 12, "Debt Outstanding", F_NORMAL)
    _c(ws, 37, 14, "=N36", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    if len(years) > 1:
        _c(ws, 37, 15, "=O36+N37", F_NORMAL, fmt=SAR_FMT, align=CENTER)
    _c(ws, 38, 12, "Debt Repayment", F_NORMAL)
    if len(years) >= 3:
        last_col = get_column_letter(13 + len(years))
        prev_col = get_column_letter(12 + len(years))
        _c(ws, 38, 13 + len(years), f"={prev_col}37", F_NORMAL, fmt=SAR_FMT, align=CENTER)

    # Net cashflows
    _c(ws, 40, 12, L.get("cf_net_cash", "Net Cashflows"), F_BOLD, border=BORDER_TOP_MED)
    for i in range(len(years)):
        _c(ws, 40, 14 + i, cf_val("net_cash_flow", i), F_BOLD, fmt=SAR_FMT, align=CENTER, border=BORDER_TOP_MED)

    # Net equity cashflow for IRR
    equity_cf = cf.get("equity_cf_for_irr", [])
    _c(ws, 41, 12, L.get("cf_net_equity", "Net Equity CF"), F_BOLD)
    _c(ws, 41, 13, equity_cf[0] if equity_cf else 0, F_BOLD, fmt=SAR_FMT, align=CENTER)
    for i in range(len(years)):
        val = equity_cf[i + 1] if i + 1 < len(equity_cf) else 0
        _c(ws, 41, 14 + i, val, F_BOLD, fmt=SAR_FMT, align=CENTER)

    # KPIs with LIVE Excel formulas
    last_yr_col = get_column_letter(13 + len(years))
    _c(ws, 43, 12, L.get("kpi_title", "Project KPIs"), F_HEADER, border=BORDER_TOP_MED)
    _c(ws, 44, 12, "IRR", F_KPI)
    _c(ws, 44, 14, f"=IRR(M41:{last_yr_col}41)", F_KPI, fmt=PCT2_FMT, align=CENTER)
    _c(ws, 45, 12, L.get("net_profit", "Net Profit"), F_BOLD)
    _c(ws, 45, 14, f"=SUM(M41:{last_yr_col}41)", F_BOLD, fmt=SAR_FMT, align=CENTER)
    _c(ws, 46, 12, L.get("roe", "ROE"), F_KPI)
    _c(ws, 46, 14, "=-N45/M41", F_KPI, fmt=PCTD_FMT, align=CENTER)
    _c(ws, 47, 12, L.get("roe_annual", "ROE Annualized"), F_BOLD)
    _c(ws, 47, 14, f"=N46/{n}", F_BOLD, fmt=PCT2_FMT, align=CENTER)

    # Data fill for input cells
    for row in range(11, 55):
        for col in [5, 6, 7, 8]:
            cell = ws.cell(row=row, column=col)
            if cell.value is not None and cell.fill == PatternFill():
                cell.fill = FILL_DATA


# ---------------------------------------------------------------------------
# Sheet 2: Zoning Report
# ---------------------------------------------------------------------------

def _build_zoning_sheet(wb: Workbook, land: dict, pf: dict, L: dict, rtl: bool) -> None:
    sheet_name = "\u062a\u0642\u0631\u064a\u0631 \u0627\u0644\u0623\u0646\u0638\u0645\u0629" if rtl else "Zoning Report"
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.rightToLeft = rtl
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["D"].width = 35

    _c(ws, 1, 2, L["zoning_title"], Font(name=FONT_AR, size=16, bold=True, color="FFFFFF"), FILL_HEADER, align=CENTER)
    ws.merge_cells("B1:D1")

    regs = land.get("regulations", {})
    plan_info = land.get("plan_info", {})
    demo = land.get("district_demographics", {})
    # Resolved area from proforma (document area overrides geoportal)
    pf_area = pf.get("inputs_used", {}).get("land_area_sqm", {}).get("value")

    items = [
        (L["parcel_no"], land.get("parcel_number")),
        (L["plan_no"], land.get("plan_number")),
        (L["district"], land.get("district_name")),
        (L["municipality"], land.get("municipality")),
        (L["area_label"], pf_area or land.get("area_sqm")),
        (L["building_code"], land.get("building_code_label")),
        ("", ""),
        (L["max_floors"], regs.get("max_floors")),
        (L["far_label"], regs.get("far")),
        (L["coverage_label"], regs.get("coverage_ratio")),
        (L["allowed_uses"], ", ".join(regs.get("allowed_uses") or [])),
        (L["setbacks_label"], ", ".join(str(v) for v in (regs.get("setback_values_m") or []))),
        (L["primary_use"], land.get("primary_use_label")),
        (L["land_use"], land.get("detailed_use_label")),
    ]

    # Plan info from Geoportal Layer 3
    if plan_info:
        items.append(("", ""))
        items.append(("\u062d\u0627\u0644\u0629 \u0627\u0644\u0645\u062e\u0637\u0637" if rtl else "Plan Status", plan_info.get("plan_status")))
        items.append(("\u0627\u0633\u062a\u062e\u062f\u0627\u0645 \u0627\u0644\u0645\u062e\u0637\u0637" if rtl else "Plan Use", plan_info.get("plan_use")))
        items.append(("\u0646\u0648\u0639 \u0627\u0644\u0645\u062e\u0637\u0637" if rtl else "Plan Type", plan_info.get("plan_type")))
        items.append(("\u062a\u0627\u0631\u064a\u062e \u0627\u0644\u0645\u062e\u0637\u0637" if rtl else "Plan Date (Hijri)", plan_info.get("plan_date_hijri")))

    # District demographics from Geoportal Layer 4
    if demo:
        items.append(("", ""))
        pop = demo.get("population")
        area_m2 = demo.get("area_m2")
        pop_str = f"{int(pop):,}" if pop else None
        items.append(("\u0633\u0643\u0627\u0646 \u0627\u0644\u062d\u064a" if rtl else "District Population", pop_str))
        if area_m2:
            area_km2 = float(area_m2) / 1_000_000
            items.append(("\u0645\u0633\u0627\u062d\u0629 \u0627\u0644\u062d\u064a" if rtl else "District Area", f"{area_km2:.1f} \u0643\u0645\u00b2"))
        en_name = demo.get("district_name_en")
        if en_name:
            items.append(("District Name (EN)", en_name))
    r = 3
    for label, val in items:
        if not label:
            r += 1
            continue
        _c(ws, r, 2, label, F_BOLD, FILL_DATA, align=RIGHT_AL, border=BORDER_THIN)
        cell = _c(ws, r, 4, val, F_NORMAL, align=CENTER, border=BORDER_THIN)
        if isinstance(val, float) and val < 1:
            cell.number_format = PCT2_FMT
        r += 1

    # Notes
    notes = regs.get("notes") or []
    if notes:
        r += 1
        _c(ws, r, 2, L["notes"], F_HEADER)
        r += 1
        for note in notes:
            # Sanitize illegal characters for openpyxl
            clean = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(note))
            _c(ws, r, 2, clean, F_SMALL, align=Alignment(wrap_text=True))
            ws.merge_cells(f"B{r}:D{r}")
            r += 1


# ---------------------------------------------------------------------------
# Sheet 3: Market Data
# ---------------------------------------------------------------------------

def _build_market_sheet(wb: Workbook, land: dict, pf: dict, L: dict, rtl: bool) -> None:
    ws = wb.create_sheet("Market Data" if not rtl else "\u0628\u064a\u0627\u0646\u0627\u062a \u0627\u0644\u0633\u0648\u0642")
    ws.sheet_view.rightToLeft = rtl
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 5
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 18

    _c(ws, 1, 2, L["market_title"], Font(name=FONT_AR, size=16, bold=True, color="FFFFFF"), FILL_HEADER, align=CENTER)
    ws.merge_cells("B1:E1")

    mkt = land.get("market", {})
    items = [
        (L["market_index"], mkt.get("srem_market_index")),
        (L["market_change"], mkt.get("srem_index_change")),
        (L["daily_transactions"], mkt.get("daily_total_transactions")),
        (L["daily_value"], mkt.get("daily_total_value_sar")),
        (L["avg_price"], mkt.get("daily_avg_price_sqm")),
    ]
    r = 3
    for label, val in items:
        _c(ws, r, 2, label, F_BOLD, FILL_DATA, align=RIGHT_AL, border=BORDER_THIN)
        _c(ws, r, 4, val, F_NORMAL, fmt=SAR_FMT if isinstance(val, (int, float)) and val and val > 100 else "0.00", align=CENTER, border=BORDER_THIN)
        r += 1

    # District-specific market data
    dist_mkt = mkt.get("district", {})
    if dist_mkt:
        r += 1
        dist_label = "\u0628\u064a\u0627\u0646\u0627\u062a \u0627\u0644\u062d\u064a" if rtl else "District Market Data"
        _c(ws, r, 2, dist_label, F_HEADER)
        r += 1
        dist_items = [
            ("\u0627\u0644\u062d\u064a" if rtl else "District", dist_mkt.get("district_name")),
            ("\u0645\u062a\u0648\u0633\u0637 \u0627\u0644\u0633\u0639\u0631 / \u0645\u00b2" if rtl else "Avg Price / m\u00b2", dist_mkt.get("avg_price_sqm")),
            ("\u0627\u0644\u0641\u062a\u0631\u0629" if rtl else "Period", dist_mkt.get("period")),
            ("\u0639\u062f\u062f \u0627\u0644\u0635\u0641\u0642\u0627\u062a" if rtl else "Total Deals", dist_mkt.get("total_deals")),
            ("\u0645\u062a\u0648\u0633\u0637 \u0627\u0644\u0631\u064a\u0627\u0636" if rtl else "Riyadh City Avg", dist_mkt.get("city_avg_price_sqm")),
        ]
        for label, val in dist_items:
            if val is not None:
                _c(ws, r, 2, label, F_BOLD, FILL_DATA, align=RIGHT_AL, border=BORDER_THIN)
                fmt = SAR_FMT if isinstance(val, (int, float)) and val and val > 100 else None
                _c(ws, r, 4, val, F_NORMAL, fmt=fmt, align=CENTER, border=BORDER_THIN)
                r += 1

    # Index history
    idx_hist = dist_mkt.get("index_history", []) if dist_mkt else []
    if idx_hist:
        r += 1
        _c(ws, r, 2, "\u0645\u0624\u0634\u0631 \u0627\u0644\u0633\u0648\u0642 (\u0623\u0633\u0628\u0648\u0639\u064a)" if rtl else "Weekly Market Index", F_HEADER)
        r += 1
        _c(ws, r, 2, "\u0627\u0644\u062a\u0627\u0631\u064a\u062e" if rtl else "Date", F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        _c(ws, r, 4, "\u0627\u0644\u0645\u0624\u0634\u0631" if rtl else "Index", F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        _c(ws, r, 5, "\u0627\u0644\u062a\u063a\u064a\u0631" if rtl else "Change", F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        r += 1
        for pt in idx_hist:
            _c(ws, r, 2, pt.get("date", ""), F_NORMAL, align=CENTER, border=BORDER_THIN)
            _c(ws, r, 4, pt.get("index"), F_NORMAL, fmt=SAR_FMT, align=CENTER, border=BORDER_THIN)
            change = pt.get("change", 0)
            cell = _c(ws, r, 5, change, F_NORMAL, fmt="0.00", align=CENTER, border=BORDER_THIN)
            if change and change >= 0:
                cell.fill = FILL_GREEN
            elif change and change < 0:
                cell.fill = FILL_RED
            r += 1

    # Intelligence KPIs
    kpis = pf.get("kpis", {})
    if kpis.get("deal_score") is not None:
        r += 1
        _c(ws, r, 2, "\u062a\u062d\u0644\u064a\u0644 \u0627\u0644\u0635\u0641\u0642\u0629" if rtl else "Deal Analysis", F_HEADER)
        r += 1
        intel_items = [
            ("\u062a\u0642\u064a\u064a\u0645 \u0627\u0644\u0635\u0641\u0642\u0629" if rtl else "Deal Score", f"{kpis['deal_score']}/100"),
            ("\u0646\u0642\u0637\u0629 \u0627\u0644\u062a\u0639\u0627\u062f\u0644" if rtl else "Break-even Price", f"{kpis.get('break_even_price_sqm', 0):,.0f} \u0631.\u0633/\u0645\u00b2"),
            ("\u062a\u0643\u0644\u0641\u0629 \u0627\u0644\u0623\u0631\u0636/\u0645\u00b2 \u0645\u0628\u0646\u064a" if rtl else "Land Cost per GBA", f"{kpis.get('land_cost_per_gba', 0):,.0f} \u0631.\u0633"),
            ("\u0645\u0636\u0627\u0639\u0641 \u0627\u0644\u0639\u0627\u0626\u062f" if rtl else "Revenue Multiple", f"{kpis.get('revenue_multiple', 0):.2f}x"),
            ("\u0646\u0633\u0628\u0629 \u0631\u0633\u0648\u0645 \u0627\u0644\u0635\u0646\u062f\u0648\u0642" if rtl else "Fund Overhead", f"{kpis.get('fund_overhead_ratio', 0)*100:.1f}%"),
        ]
        for label, val in intel_items:
            _c(ws, r, 2, label, F_BOLD, FILL_DATA, align=RIGHT_AL, border=BORDER_THIN)
            _c(ws, r, 4, val, F_NORMAL, align=CENTER, border=BORDER_THIN)
            r += 1

        # Risk flags
        risks = kpis.get("risk_flags", [])
        if risks:
            r += 1
            _c(ws, r, 2, "\u0645\u062e\u0627\u0637\u0631" if rtl else "Risk Flags", F_BOLD)
            r += 1
            for flag in risks:
                _c(ws, r, 2, flag, F_NORMAL, FILL_RED, border=BORDER_THIN)
                r += 1

    # Trending districts
    trending = mkt.get("trending_districts") or []
    if trending:
        r += 2
        _c(ws, r, 2, L["trending_title"], F_HEADER)
        r += 1
        _c(ws, r, 2, L["district"], F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        _c(ws, r, 3, L["city"], F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        _c(ws, r, 4, L["deals"], F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        _c(ws, r, 5, L["value_sar"], F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
        r += 1
        for d in trending:
            _c(ws, r, 2, d.get("name"), F_NORMAL, align=CENTER, border=BORDER_THIN)
            _c(ws, r, 3, d.get("city"), F_NORMAL, align=CENTER, border=BORDER_THIN)
            _c(ws, r, 4, d.get("deals"), F_NORMAL, align=CENTER, border=BORDER_THIN)
            _c(ws, r, 5, d.get("total_sar"), F_NORMAL, fmt=SAR_FMT, align=CENTER, border=BORDER_THIN)
            r += 1

    # Data sources
    sources = land.get("data_sources", {})
    if sources:
        r += 2
        _c(ws, r, 2, "\u0645\u0635\u0627\u062f\u0631 \u0627\u0644\u0628\u064a\u0627\u0646\u0627\u062a" if rtl else "Data Sources", F_HEADER)
        r += 1
        for src, active in sources.items():
            status = "\u2705 \u0646\u0634\u0637" if active else "\u274c \u063a\u064a\u0631 \u0645\u062a\u0627\u062d"
            _c(ws, r, 2, src.replace("_", " "), F_NORMAL, border=BORDER_THIN)
            cell = _c(ws, r, 4, status, F_NORMAL, align=CENTER, border=BORDER_THIN)
            cell.fill = FILL_GREEN if active else FILL_RED
            r += 1

    r += 2
    _c(ws, r, 2, L["market_source"], F_SMALL)


# ---------------------------------------------------------------------------
# Sheet 4: Sensitivity
# ---------------------------------------------------------------------------

def _build_sensitivity_sheet(wb: Workbook, pf: dict, L: dict, rtl: bool) -> None:
    ws = wb.create_sheet("Sensitivity" if not rtl else "تحليل الحساسية")
    ws.sheet_view.rightToLeft = rtl

    sens = pf.get("sensitivity")
    if not sens:
        _c(ws, 2, 2, "لا تتوفر بيانات", F_NORMAL)
        return

    _c(ws, 1, 2, "تحليل الحساسية - معدل العائد الداخلي (IRR)", Font(name=FONT_AR, size=16, bold=True, color="FFFFFF"), FILL_HEADER, align=CENTER)
    ws.merge_cells("B1:G1")

    sale_range = sens["sale_price_range"]
    cost_range = sens["construction_cost_range"]
    matrix = sens["irr_matrix"]

    ws.column_dimensions["B"].width = 18
    for i in range(len(cost_range)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    _c(ws, 3, 2, "سعر البيع ↓ \\ التكلفة →", F_BOLD, align=CENTER)
    for i, c in enumerate(cost_range):
        _c(ws, 3, 3 + i, c, F_BOLD, fmt="#,##0", align=CENTER, border=BORDER_THIN)

    for ri, sp in enumerate(sale_range):
        r = 4 + ri
        _c(ws, r, 2, sp, F_BOLD, fmt="#,##0", align=CENTER, border=BORDER_THIN)
        for ci, irr in enumerate(matrix[ri]):
            cell = _c(ws, r, 3 + ci, irr, F_NORMAL, fmt=PCT2_FMT, align=CENTER, border=BORDER_THIN)
            if irr is not None:
                if irr >= 0.10:
                    cell.fill = FILL_GREEN
                elif irr >= 0:
                    cell.fill = FILL_YELLOW
                else:
                    cell.fill = FILL_RED


# ---------------------------------------------------------------------------
# Sheet 5: Scenario Comparison
# ---------------------------------------------------------------------------

def _build_scenario_sheet(
    wb: Workbook, scenario_results: list[dict], L: dict, rtl: bool,
) -> None:
    """Scenario comparison with fully computed numbers for all 3 scenarios."""
    ws = wb.create_sheet("\u0645\u0642\u0627\u0631\u0646\u0629" if rtl else "Scenarios")
    ws.sheet_view.rightToLeft = rtl
    ws.column_dimensions["B"].width = 28
    for c in ["C", "D", "E"]:
        ws.column_dimensions[c].width = 22

    title = "\u0645\u0642\u0627\u0631\u0646\u0629 \u0627\u0644\u0633\u064a\u0646\u0627\u0631\u064a\u0648\u0647\u0627\u062a" if rtl else "Scenario Comparison"
    _c(ws, 1, 2, title, Font(name=FONT_AR, size=16, bold=True, color="FFFFFF"), FILL_HEADER, align=CENTER)
    ws.merge_cells("B1:E1")

    names = [
        "\u0645\u062a\u062d\u0641\u0638" if rtl else "Conservative",
        "\u0623\u0633\u0627\u0633\u064a" if rtl else "Base",
        "\u062c\u0631\u064a\u0621" if rtl else "Aggressive",
    ]

    _c(ws, 3, 2, "\u0627\u0644\u0633\u064a\u0646\u0627\u0631\u064a\u0648" if rtl else "Scenario", F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)
    for i, name in enumerate(names):
        _c(ws, 3, 3 + i, name, F_BOLD, FILL_SECTION, align=CENTER, border=BORDER_THIN)

    # Input assumptions row
    assumption_rows = [
        ("\u0633\u0639\u0631 \u0627\u0644\u0628\u064a\u0639 / \u0645\u00b2" if rtl else "Sale Price / m\u00b2", "sale_price_per_sqm", SAR_FMT),
        ("\u062a\u0643\u0644\u0641\u0629 \u0627\u0644\u0628\u0646\u064a\u0629 \u0627\u0644\u062a\u062d\u062a\u064a\u0629" if rtl else "Infrastructure / m\u00b2", "infrastructure_cost_per_sqm", SAR_FMT),
        ("\u062a\u0643\u0644\u0641\u0629 \u0627\u0644\u0628\u0646\u064a\u0629 \u0627\u0644\u0639\u0644\u0648\u064a\u0629" if rtl else "Superstructure / m\u00b2", "superstructure_cost_per_sqm", SAR_FMT),
    ]
    r = 4
    for label, key, fmt in assumption_rows:
        _c(ws, r, 2, label, F_NORMAL, align=CENTER, border=BORDER_THIN)
        for i, sc in enumerate(scenario_results):
            val = sc.get("inputs_used", {}).get(key, {}).get("value", 0)
            _c(ws, r, 3 + i, val, F_NORMAL, fmt=fmt, align=CENTER, border=BORDER_THIN)
        r += 1

    r += 1  # blank row

    # KPI rows — fully computed
    kpi_rows = [
        ("\u062d\u062c\u0645 \u0627\u0644\u0635\u0646\u062f\u0648\u0642" if rtl else "Fund Size", "fund_size", "total_fund_size", SAR_FMT),
        ("\u062d\u0642\u0648\u0642 \u0627\u0644\u0645\u0644\u0643\u064a\u0629" if rtl else "Equity", "fund_size", "equity_amount", SAR_FMT),
        ("\u0625\u062c\u0645\u0627\u0644\u064a \u0627\u0644\u0625\u064a\u0631\u0627\u062f\u0627\u062a" if rtl else "Revenue", "revenue", "gross_revenue", SAR_FMT),
        ("\u0635\u0627\u0641\u064a \u0627\u0644\u0631\u0628\u062d" if rtl else "Net Profit", "kpis", "equity_net_profit", SAR_FMT),
        ("\u0645\u0639\u062f\u0644 \u0627\u0644\u0639\u0627\u0626\u062f \u0627\u0644\u062f\u0627\u062e\u0644\u064a" if rtl else "IRR", "kpis", "irr", PCT2_FMT),
        ("\u0627\u0644\u0639\u0627\u0626\u062f \u0639\u0644\u0649 \u0627\u0644\u0645\u0644\u0643\u064a\u0629" if rtl else "ROE", "kpis", "roe_total", PCTD_FMT),
        ("\u062a\u0642\u064a\u064a\u0645 \u0627\u0644\u0635\u0641\u0642\u0629" if rtl else "Deal Score", "kpis", "deal_score", "#,##0"),
    ]
    for label, section, key, fmt in kpi_rows:
        _c(ws, r, 2, label, F_BOLD, FILL_DATA, align=CENTER, border=BORDER_THIN)
        for i, sc in enumerate(scenario_results):
            val = sc.get(section, {}).get(key)
            cell = _c(ws, r, 3 + i, val, F_NORMAL, fmt=fmt, align=CENTER, border=BORDER_THIN)
            # Color code IRR
            if key == "irr" and val is not None:
                if val >= 0.10:
                    cell.fill = FILL_GREEN
                elif val >= 0:
                    cell.fill = FILL_YELLOW
                else:
                    cell.fill = FILL_RED
        r += 1


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel(
    result: dict,
    land_object: dict,
    overrides: dict | None = None,
    lang: str = "ar",
) -> bytes:
    """Generate a professional .xlsx with 5 sheets and live formulas.

    Args:
        result: ProFormaResult from computation_engine (base scenario).
        land_object: Land Object from data_fetch.
        overrides: User overrides dict (for computing scenarios).
        lang: 'ar' for Arabic (default), 'en' for English.
    """
    from computation_engine import compute_proforma

    wb = Workbook()
    L = _L(lang)
    rtl = (lang == "ar")

    # Compute 3 scenarios: conservative, base, aggressive
    base_ov = dict(overrides or {})
    sale = result.get("inputs_used", {}).get("sale_price_per_sqm", {}).get("value", 6000) or 6000
    infra = result.get("inputs_used", {}).get("infrastructure_cost_per_sqm", {}).get("value", 500) or 500
    super_ = result.get("inputs_used", {}).get("superstructure_cost_per_sqm", {}).get("value", 2500) or 2500

    conservative_ov = {**base_ov, "sale_price_per_sqm": sale * 0.8,
                       "infrastructure_cost_per_sqm": infra * 1.1, "superstructure_cost_per_sqm": super_ * 1.1}
    aggressive_ov = {**base_ov, "sale_price_per_sqm": sale * 1.3,
                     "infrastructure_cost_per_sqm": infra * 0.9, "superstructure_cost_per_sqm": super_ * 0.9}

    try:
        conservative = compute_proforma(land_object, conservative_ov)
    except Exception:
        conservative = result  # fallback
    try:
        aggressive = compute_proforma(land_object, aggressive_ov)
    except Exception:
        aggressive = result

    scenario_results = [conservative, result, aggressive]

    _build_assumptions_sheet(wb, result, land_object, L, rtl)
    _build_zoning_sheet(wb, land_object, result, L, rtl)
    _build_market_sheet(wb, land_object, result, L, rtl)
    _build_sensitivity_sheet(wb, result, L, rtl)
    _build_scenario_sheet(wb, scenario_results, L, rtl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
