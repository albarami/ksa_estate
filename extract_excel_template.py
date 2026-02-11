"""Extract the Al-Hada Excel into a generic computation template.

Reads every sheet, captures all formulas, values, labels, and structure
to produce a computation_template.json that works for ANY parcel.
"""

import json
import re
from pathlib import Path

import openpyxl

EXCEL = Path(r"C:\Projects\ksa_estate\Al-Hada Opportunity.xlsx")
OUT = Path("al_hada_excel_extract.json")

wb = openpyxl.load_workbook(str(EXCEL), data_only=False)
wb_values = openpyxl.load_workbook(str(EXCEL), data_only=True)

result: dict = {
    "sheet_names": wb.sheetnames,
    "sheets": {},
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_val = wb_values[sheet_name]
    
    sheet_data: dict = {
        "dimensions": ws.dimensions,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "merged_cells": [str(m) for m in ws.merged_cells.ranges],
        "rows": [],
    }
    
    for row_idx in range(1, min(ws.max_row + 1, 200)):
        row_data: list = []
        for col_idx in range(1, min(ws.max_column + 1, 30)):
            cell = ws.cell(row=row_idx, column=col_idx)
            val_cell = ws_val.cell(row=row_idx, column=col_idx)
            
            cell_info: dict = {}
            
            if cell.value is not None:
                cell_info["formula"] = str(cell.value) if str(cell.value).startswith("=") else None
                cell_info["value"] = val_cell.value
                cell_info["type"] = type(val_cell.value).__name__ if val_cell.value is not None else "None"
                
                # Convert to serializable
                if cell_info["value"] is not None:
                    try:
                        json.dumps(cell_info["value"])
                    except (TypeError, ValueError):
                        cell_info["value"] = str(cell_info["value"])
                
                if cell.number_format and cell.number_format != "General":
                    cell_info["format"] = cell.number_format
                
                row_data.append({
                    "col": col_idx,
                    "ref": cell.coordinate,
                    **cell_info,
                })
        
        if row_data:
            sheet_data["rows"].append({
                "row": row_idx,
                "cells": row_data,
            })
    
    result["sheets"][sheet_name] = sheet_data

OUT.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
print(f"Extracted {len(result['sheets'])} sheets to {OUT}")
for name, data in result["sheets"].items():
    n_rows = len(data["rows"])
    print(f"  {name}: {n_rows} rows, {data['max_col']} cols, {len(data['merged_cells'])} merges")

wb.close()
wb_values.close()
